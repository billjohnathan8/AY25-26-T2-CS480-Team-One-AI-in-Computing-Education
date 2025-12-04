from __future__ import annotations

import csv
import io
import json
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

from ..schemas import CourseCreate, StudentCreate, SubmissionCreate


def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _load_rows(file_bytes: bytes, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".csv", ".txt"}:
        reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8")))
        return [row for row in reader]
    if suffix in {".xlsx", ".xls"}:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheet = workbook.active
        rows = []
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(max_row=1))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({headers[idx]: row[idx] for idx in range(len(headers))})
        return rows
    raise ValueError("Unsupported file type for tabular import")


def parse_courses_from_file(file_bytes: bytes, filename: str) -> list[CourseCreate]:
    rows = _load_rows(file_bytes, filename)
    records: list[CourseCreate] = []
    for row in rows:
        manifest = row.get("content_manifest")
        if isinstance(manifest, (dict, list)):
            manifest = json.dumps(manifest)
        records.append(
            CourseCreate(
                name=str(_clean(row.get("name") or row.get("course_name") or "")),
                section_number=_clean(row.get("section_number")),
                description=_clean(row.get("description")),
                content_manifest=manifest,
            )
        )
    return records


def parse_students_from_file(file_bytes: bytes, filename: str) -> list[StudentCreate]:
    rows = _load_rows(file_bytes, filename)
    students: list[StudentCreate] = []
    for row in rows:
        students.append(
            StudentCreate(
                name=str(_clean(row.get("name") or row.get("student_name") or "")),
                email=str(_clean(row.get("email") or row.get("student_email") or "")),
            )
        )
    return students


def extract_pdf_text(file_bytes: bytes) -> str:
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        contents = []
        for page in pdf.pages:
            contents.append(page.extract_text() or "")
        return "\n".join(contents).strip()


def parse_submissions_file(
    file_bytes: bytes,
    filename: str,
    default_course_id: int | None = None,
    default_student_email: str | None = None,
) -> list[SubmissionCreate]:
    suffix = Path(filename).suffix.lower()
    submissions: list[SubmissionCreate] = []
    if suffix == ".pdf":
        text = extract_pdf_text(file_bytes)
        submissions.append(
            SubmissionCreate(
                course_id=default_course_id,
                student_email=default_student_email,
                answer_text=text,
                source_filename=filename,
                ocr_text=text,
            )
        )
        return submissions

    for row in _load_rows(file_bytes, filename):
        submissions.append(
            SubmissionCreate(
                student_email=_clean(row.get("student_email") or row.get("email")) or default_student_email,
                student_id=_clean(row.get("student_id")),
                course_id=_clean(row.get("course_id")) or default_course_id,
                course_name=_clean(row.get("course_name")),
                topic_title=_clean(row.get("topic_title") or row.get("topic")),
                topic_category=_clean(row.get("topic_category")),
                answer_text=str(_clean(row.get("answer_text") or row.get("answer") or "")),
                raw_score=_clean(row.get("raw_score")),
                final_score=_clean(row.get("final_score")),
                exam_type=_clean(row.get("exam_type") or row.get("exam")),
                source_filename=filename,
            )
        )
    return submissions
